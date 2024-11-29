#!/usr/bin/env python
# coding: utf-8

# # pip package

# In[1]:


# !pip install -r multimodal/env/requirement.txt
# !pip3 uninstall --yes torch torchaudio torchvision torchtext torchdata
# !pip3 install torch torchaudio torchvision torchtext torchdata
# !pip install torch_tb_profiler


# # The imports
# 

# In[2]:


import sys
from datetime import datetime
# from torch.nn import LayerNorm, Linear, Dropout, Softmax
from einops import rearrange, repeat
import copy
from timm.models.layers import DropPath, trunc_normal_
import re
import torch.backends.cudnn as cudnn

import matplotlib.pyplot as plt
# from torchsummary import summary
from sklearn.metrics import confusion_matrix, accuracy_score, classification_report, cohen_kappa_score
from operator import truediv
import math
from PIL import Image
import time
import torchvision.transforms.functional as TF
from torch.nn.parameter import Parameter
from sklearn.decomposition import PCA
from scipy.io import loadmat as loadmat
from scipy import io
import torch.utils.data as dataf
import torch.nn as nn
import torch
import torch.nn.functional as F
from torch import einsum
import random
import numpy as np
import os
import torchvision
from pathlib import Path
from openpyxl import load_workbook,Workbook
cudnn.deterministic = True
cudnn.benchmark = False
os.environ['CUDA_LAUNCH_BLOCKING'] = "1"
from torch.utils.tensorboard import SummaryWriter
from pathlib import Path
import configparser
from matplotlib import colors


# In[ ]:





# In[3]:


# # 确定自己是否在GPU环境下，如果输出gpu则证明在

# 查看显卡驱动
# !nvidia-smi

print(torch.cuda.device_count())
print(torch.cuda.is_available())



# In[4]:


# !kill 5328
# %reload_ext tensorboard
# %load_ext tensorboard
# %tensorboard --logdir /root/autodl-tmp/multimodal/runs


# # Choose the param Config

# In[5]:


# # # jupyter需要下面一行
# sys.argv = ['multimodal.py ']

import argparse
parser = argparse.ArgumentParser()

parser.add_argument("--configName", help="sectionName", type=str,default="TEST")

args = parser.parse_args()

configName = args.configName

current_time = datetime.now().strftime('%b%d_%H-%M-%S')

config = configparser.ConfigParser()
config.read('paramConfig.ini')

section = config[configName] 

# config of dataset names and sources
modalName1 = section["modalName1"]
modalName2 = section["modalName2"]
datasetName = section["datasetName"]
# config of training params
patchsize = int(section["patchsize"])
batchsize = int(section["batchsize"])
testSizeNumber = int(section["testSizeNumber"])
EPOCH = int(section["EPOCH"])
LR = float(section["LR"])
HSIOnly = section.getboolean("HSIOnly")
num_workers = int(section["num_workers"])

# config of paths
checkpointName = section['checkpointName']
parent_directory = Path(section["parent_directory"])
if section.getboolean("self_dataset"):
    datasetPath = parent_directory / "dataset" / "img"
else:
    datasetPath = parent_directory / "dataset"
checkpointPath = parent_directory / "checkpoint"
resultPath = parent_directory / "result"
imgPath = parent_directory / "visresult"
tensorboardPath = parent_directory / "runs" / f"{current_time}_{checkpointName}_{configName}"
checkpointDatasetPath = checkpointPath / datasetName
datasetPath.mkdir(parents=True, exist_ok=True)
checkpointPath.mkdir(parents=True, exist_ok=True)
resultPath.mkdir(parents=True, exist_ok=True)
imgPath.mkdir(parents=True, exist_ok=True)
checkpointDatasetPath.mkdir(parents=True, exist_ok=True)

writer = SummaryWriter(log_dir=tensorboardPath )


# In[6]:


params =  dict(section.items())
params['configName']=configName
params


# # The DatasetConfig Class Define

# In[7]:


class DatasetConfig:
    def __init__(self,name):
        self.name = name
        self.classNum = None
        self.bandNum1 = None
        self.bandNum2 = None
        self.shape1 = None
        self.shape2 = None
        self.path = datasetPath / self.name


    def getTeOrTr(self,dataSourceName,TrOrTe="Tr"):
        # read the modal data
        trPath = self.path / "{}_{}.mat".format(dataSourceName,TrOrTe)
        trainPatch = loadmat(trPath)['Data'].astype(np.float32)
        return trainPatch
    def getLabel(self,TrOrTe="Tr"):
        # read the modal data
        labelPath = self.path / '{}Label.mat'.format(TrOrTe)
        label = loadmat(labelPath)['Data']
        return label

    def getTrainLoader(self,type = "Tr",batchsize = batchsize):
        # HSI patch
        TrainPatch1 = self.getTeOrTr(modalName1,type)#(2832, 11, 11, 144)
        TrainPatch1 = torch.from_numpy(TrainPatch1).to(torch.float32)
        TrainPatch1 = TrainPatch1.permute(0,3,1,2)
        TrainPatch1 = TrainPatch1.reshape(TrainPatch1.shape[0],TrainPatch1.shape[1],-1).to(torch.float32) # 2832, 144, 121

        # LIDAR patch
        TrainPatch2 = self.getTeOrTr(modalName2,type)#(12197, 11, 11, 144)
        TrainPatch2 = torch.from_numpy(TrainPatch2).to(torch.float32)
        TrainPatch2 = TrainPatch2.permute(0,3,1,2)
        TrainPatch2 = TrainPatch2.reshape(TrainPatch2.shape[0],TrainPatch2.shape[1],-1).to(torch.float32)

        # Label
        TrainLabel1 = self.getLabel(type)#(2832, 11, 11, 144)
        TrainLabel1 = torch.from_numpy(TrainLabel1)-1
        TrainLabel1 = TrainLabel1.long()
        TrainLabel1 = TrainLabel1.reshape(-1)

        print("{} {} data shape = {}".format(modalName1,type,TrainPatch1.shape) )
        print("{} {} data shape = {}".format(modalName2,type,TrainPatch2.shape))
        print("{} label shape = {}".format(type,TrainLabel1.shape))
        dataset = dataf.TensorDataset(TrainPatch1,TrainPatch2, TrainLabel1)
        trainLoader = dataf.DataLoader(dataset, batch_size=batchsize, shuffle=True, num_workers= num_workers)
        #record class number
        if self.classNum==None:
            self.classNum = len(np.unique(TrainLabel1))
        #record HSI band number
        if self.bandNum1==None:
            self.bandNum1 = TrainPatch1.shape[1]
        #record LIDAR band number
        if self.bandNum2==None:
            self.bandNum2 = TrainPatch2.shape[1]
        #record HSI shape
        if self.shape1==None:
            self.shape1 = TrainPatch1.shape
        #record LiDAR shape
        if self.shape2==None:
            self.shape2 = TrainPatch2.shape

        return trainLoader






# # Train tool

# In[8]:


def getResolution(text):
    """Matches a string of the format (number, number) and returns a tuple"""
    pattern = r'\(([0-9]+),([0-9]+),([0-9]+)\)'
    match = re.match(pattern,text)
    resolution = tuple((int(x) for x in match.groups()))     
    return resolution   

def AA_andEachClassAccuracy(confusion_matrix):
    counter = confusion_matrix.shape[0]
    list_diag = np.diag(confusion_matrix)
    list_raw_sum = np.sum(confusion_matrix, axis=1)
    each_acc = np.nan_to_num(truediv(list_diag, list_raw_sum))
    average_acc = np.mean(each_acc)
    return each_acc, average_acc

def recordExcel(oa,aa,kappa,other,each_acc):
    rows = [oa,aa,kappa,other,current_time,configName,section['network']]
    for each in each_acc[0]:
        rows.append(each)
    
    sheetName = datasetName+"_"+modalName2 if (modalName2=='MS' or  modalName2=='SAR' or  modalName2=='DSM') else datasetName
    # Excel文件路径
    excel_file_path = resultPath / "expRes.xlsx"
    try:
        # 打开Excel文件
        workbook = load_workbook(excel_file_path)
        sheet = workbook[sheetName]
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheetName
        sheet.append(["OA","AA","Kappa","Other","Time","Config","Network"])
    except KeyError:
        sheet = workbook.create_sheet(title=sheetName)
        sheet.append(["OA","AA","Kappa","Other","Time","Config","Network"])
    # 将一行内容追加到工作表
    sheet.append(rows)
    # 保存修改后的Excel文件
    workbook.save(excel_file_path)

    print(f"write finished: {excel_file_path}")

def reports (testloader,model,classNum,name="Houston"):
    if name == 'Houston':
        target_names = ['Healthy grass', 'Stressed grass', 'Synthetic grass'
                        ,'Trees', 'Soil', 'Water',
                        'Residential', 'Commercial', 'Road', 'Highway',
                        'Railway', 'Parking Lot 1', 'Parking Lot 2', 'Tennis Court',
                        'Running Track']
    elif name == 'Trento':
        target_names = ['Apples','Buildings','Ground','Woods','Vineyard',
                        'Roads']
    elif name == 'MUUFL' or name == 'MUUFLS' or name == 'MUUFLSR':
        target_names = ['Trees','Grass_Pure','Grass_Groundsurface','Dirt_And_Sand', 'Road_Materials','Water',"Buildings'_Shadow",
                    'Buildings','Sidewalk','Yellow_Curb','ClothPanels']
    elif name == 'Augsburg':
        target_names =  ['Forest','Residential-Area','Industrail-Area','Low-Plants','Allotment','Commericial-Area','Water']
    elif name == 'IP':
        target_names = ['Alfalfa', 'Corn-notill', 'Corn-mintill', 'Corn'
                ,'Grass-pasture', 'Grass-trees', 'Grass-pasture-mowed',
                'Hay-windrowed', 'Oats', 'Soybean-notill', 'Soybean-mintill',
                'Soybean-clean', 'Wheat', 'Woods', 'Buildings-Grass-Trees-Drives',
                'Stone-Steel-Towers']
    elif name == 'SA':
        target_names = ['Brocoli_green_weeds_1','Brocoli_green_weeds_2','Fallow','Fallow_rough_plow','Fallow_smooth',
                        'Stubble','Celery','Grapes_untrained','Soil_vinyard_develop','Corn_senesced_green_weeds',
                        'Lettuce_romaine_4wk','Lettuce_romaine_5wk','Lettuce_romaine_6wk','Lettuce_romaine_7wk',
                        'Vinyard_untrained','Vinyard_vertical_trellis']
    elif name == 'UP':
        target_names = ['Asphalt','Meadows','Gravel','Trees', 'Painted metal sheets','Bare Soil','Bitumen',
                        'Self-Blocking Bricks','Shadows']
    tar,pre = valid(testloader,model)
    print(tar,pre)
    oa = accuracy_score(tar, pre)
    confusion = confusion_matrix(tar, pre,labels=range(classNum))
    print(confusion)
    each_acc, aa = AA_andEachClassAccuracy(confusion)
    kappa = cohen_kappa_score(tar, pre)
    print("Final result:")
    print("OA: {:.4f} | AA: {:.4f} | Kappa: {:.4f}".format(oa, aa, kappa))
    print(each_acc)
    print("**************************************************")
    return confusion, oa*100, each_acc*100, aa*100, kappa*100

def set_seed(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    np.random.seed(seed)


def save_checkpoint(folder_path,fileName,model):
    # save the checkpoint
    fileName = folder_path / fileName
    torch.save(model.state_dict(),fileName)
    return fileName

def record_output(oa_ae, aa_ae, kappa_ae, element_acc_ae, path):
    f = open(path, 'w')
    sentence0 = 'OAs for each iteration are:' + str(oa_ae) + '\n'
    f.write(sentence0)
    sentence1 = 'AAs for each iteration are:' + str(aa_ae) + '\n'
    f.write(sentence1)
    sentence2 = 'KAPPAs for each iteration are:' + str(kappa_ae) + '\n' + '\n'
    f.write(sentence2)
    sentence2_1 = 'each acc  is: ' + str(element_acc_ae) + '\n'
    f.write(sentence2_1)
    sentence3 = 'mean_OA ± std_OA is: ' + str(np.mean(oa_ae)) + ' ± ' + str(np.std(oa_ae)) + '\n'
    f.write(sentence3)
    sentence4 = 'mean_AA ± std_AA is: ' + str(np.mean(aa_ae)) + ' ± ' + str(np.std(aa_ae)) + '\n'
    f.write(sentence4)
    sentence5 = 'mean_KAPPA ± std_KAPPA is: ' + str(np.mean(kappa_ae)) + ' ± ' + str(np.std(kappa_ae)) + '\n' + '\n'
    f.write(sentence5)

    element_mean = np.mean(element_acc_ae, axis=0)
    element_std = np.std(element_acc_ae, axis=0)
    sentence8 = "Mean of all elements in confusion matrix: " + str(element_mean) + '\n'
    f.write(sentence8)
    sentence9 = "Standard deviation of all elements in confusion matrix: " + str(element_std) + '\n' + '\n'
    f.write(sentence9)
    element_mean = list(element_mean)
    element_mean.extend([np.mean(oa_ae),np.mean(aa_ae),np.mean(kappa_ae)])
    element_std = list(element_std)
    element_std.extend([np.std(oa_ae),np.std(aa_ae),np.std(kappa_ae)])
    sentence10 = "All values without std: " + str(element_mean) + '\n' + '\n'
    f.write(sentence10)
    sentence11 = "All values with std: "
    for i,x in enumerate(element_mean):
        sentence11 += str(element_mean[i]) + " ± " +  str(element_std[i]) + ", "
    sentence11 += "\n"
    f.write(sentence11)
    f.write(str(params))
    
    f.close()


# # Create Model

# In[9]:


#MFT WITH CHANNEL TOKENIZATION

from torch.nn import LayerNorm,Linear,Dropout,Softmax
import copy



def INF(B,H,W):
     return -torch.diag(torch.tensor(float("inf")).cuda().repeat(H),0).unsqueeze(0).repeat(B*W,1,1)

     
class HetConv(nn.Module):
    def __init__(self, in_channels, out_channels, kernel_size=3, stride=1,padding = None, bias = None,p = 64, g = 64):
        super(HetConv, self).__init__()
        # Groupwise Convolution
        self.gwc = nn.Conv2d(in_channels, out_channels, kernel_size=kernel_size,groups=g,padding = kernel_size//3, stride = stride)
        # Pointwise Convolution
        self.pwc = nn.Conv2d(in_channels, out_channels, kernel_size=1,groups=p, stride = stride)
    def forward(self, x):
        return self.gwc(x) + self.pwc(x)   

class MCrossAttention(nn.Module):
    def __init__(self, dim, num_heads=8, qkv_bias=False, qk_scale=None, attn_drop=0.1, proj_drop=0.1):
        super().__init__()
        self.num_heads = num_heads
        head_dim = dim // num_heads
        self.scale = qk_scale or head_dim ** -0.5

        self.wq = nn.Linear(head_dim, dim , bias=qkv_bias)
        self.wk = nn.Linear(head_dim, dim , bias=qkv_bias)
        self.wv = nn.Linear(head_dim, dim , bias=qkv_bias)
#         self.attn_drop = nn.Dropout(attn_drop)
        self.proj = nn.Linear(dim * num_heads, dim)
        self.proj_drop = nn.Dropout(proj_drop)

    def forward(self, x):

        B, N, C = x.shape
        q = self.wq(x[:, 0:1, ...].reshape(B, 1, self.num_heads, C // self.num_heads)).permute(0, 2, 1, 3)  # B1C -> B1H(C/H) -> BH1(C/H)
        k = self.wk(x.reshape(B, N, self.num_heads, C // self.num_heads)).permute(0, 2, 1, 3)  # BNC -> BNH(C/H) -> BHN(C/H)
        v = self.wv(x.reshape(B, N, self.num_heads, C // self.num_heads)).permute(0, 2, 1, 3)  # BNC -> BNH(C/H) -> BHN(C/H)
        attn = torch.einsum('bhid,bhjd->bhij', q, k) * self.scale
#         attn = (q @ k.transpose(-2, -1)) * self.scale  # BH1(C/H) @ BH(C/H)N -> BH1N
        attn = attn.softmax(dim=-1)
#         attn = self.attn_drop(attn)
        x = torch.einsum('bhij,bhjd->bhid', attn, v).transpose(1, 2)
#         x = (attn @ v).transpose(1, 2)
        x = x.reshape(B, 1, C * self.num_heads)   # (BH1N @ BHN(C/H)) -> BH1(C/H) -> B1H(C/H) -> B1C
        x = self.proj(x)
        x = self.proj_drop(x)
        return x

class Mlp(nn.Module):
    def __init__(self, dim):
        super(Mlp, self).__init__()
        self.fc1 = Linear(dim, 512)
        self.fc2 = Linear(512, dim)
        self.act_fn = nn.GELU()
        self.dropout = Dropout(0.1)

        self._init_weights()

    def _init_weights(self):
        nn.init.xavier_uniform_(self.fc1.weight)
        nn.init.xavier_uniform_(self.fc2.weight)
        nn.init.normal_(self.fc1.bias, std=1e-6)
        nn.init.normal_(self.fc2.bias, std=1e-6)

    def forward(self, x):
        x = self.fc1(x)
        x = self.act_fn(x)
        x = self.dropout(x)
        x = self.fc2(x)
        x = self.dropout(x)
        return x
class Block(nn.Module):
    def __init__(self, dim):
        super(Block, self).__init__()
        self.hidden_size = dim
        self.attention_norm = LayerNorm(dim, eps=1e-6)
        self.ffn_norm = LayerNorm(dim, eps=1e-6)
        self.ffn = Mlp(dim)
#         self.attn = Attention(dim = 64)
        self.attn = MCrossAttention(dim = dim)
    def forward(self, x):
        h = x
        x = self.attention_norm(x)
        x= self.attn(x)
        x = x + h

        h = x
        x = self.ffn_norm(x)
        x = self.ffn(x)
        x = x + h
        
        return x

class TransformerEncoder(nn.Module):

    def __init__(self, dim, num_heads= 8, mlp_ratio=4., qkv_bias=False, qk_scale=None, drop=0.1, attn_drop=0.1,
                 drop_path=0.1, act_layer=nn.GELU, norm_layer=nn.LayerNorm, has_mlp=False):
        super().__init__()
        self.layer = nn.ModuleList()
        self.encoder_norm = LayerNorm(dim, eps=1e-6)
        for _ in range(2):
            layer = Block(dim)
            self.layer.append(copy.deepcopy(layer))
       

    def forward(self, x):
        for layer_block in self.layer:
            x= layer_block(x)
            
        encoded = self.encoder_norm(x)
       
        

        return encoded[:,0]


class MFT(nn.Module):
    def __init__(self, FM, NC, NCLidar, Classes, HSIOnly,patchsize):
        super(MFT, self).__init__()
        self.HSIOnly = HSIOnly
        self.conv5 = nn.Sequential(
            nn.Conv3d(1, 8, (9, 3, 3), padding=(0,1,1), stride = 1),
            nn.BatchNorm3d(8),
            nn.ReLU()
        )
        
        self.conv6 = nn.Sequential(
            HetConv(8 * (NC - 8), FM*4,
                p = 1,
                g = (FM*4)//4 if (8 * (NC - 8))%FM == 0 else (FM*4)//8,
                   ),
            nn.BatchNorm2d(FM*4),
            nn.ReLU()
        )
        
        self.last_BandSize = NC//2//2//2
        
        self.lidarConv = nn.Sequential(
                        nn.Conv2d(NCLidar,64,3,1,1),
                        nn.BatchNorm2d(64),
                        nn.GELU()
                        )
        self.ca = TransformerEncoder(FM*4)
        self.out3 = nn.Linear(FM*4 , Classes)
        self.position_embeddings = nn.Parameter(torch.randn(1, 4 + 1, FM*4))
        self.dropout = nn.Dropout(0.1)
        torch.nn.init.xavier_uniform_(self.out3.weight)
        torch.nn.init.normal_(self.out3.bias, std=1e-6)
        self.token_wA = nn.Parameter(torch.empty(1, 4, 64),
                                     requires_grad=True)  # Tokenization parameters
        torch.nn.init.xavier_normal_(self.token_wA)
        self.token_wV = nn.Parameter(torch.empty(1, 64, 64),
                                     requires_grad=True)  # Tokenization parameters
        torch.nn.init.xavier_normal_(self.token_wV)
        
        self.token_wA_L = nn.Parameter(torch.empty(1, 1, 64),
                                     requires_grad=True)  # Tokenization parameters
        torch.nn.init.xavier_normal_(self.token_wA_L)
        self.token_wV_L = nn.Parameter(torch.empty(1, 64, 64),
                                     requires_grad=True)  # Tokenization parameters
        torch.nn.init.xavier_normal_(self.token_wV_L)
        self.patchsize = patchsize
        

    def forward(self, x1, x2):
        x1 = x1.reshape(x1.shape[0],-1,self.patchsize,self.patchsize)
        x1 = x1.unsqueeze(1)
        x2 = x2.reshape(x2.shape[0],-1,self.patchsize,self.patchsize)
        x1 = self.conv5(x1)
        x1 = x1.reshape(x1.shape[0],-1,self.patchsize,self.patchsize)
        
        x1 = self.conv6(x1)
        x2 = self.lidarConv(x2)
        x2 = x2.reshape(x2.shape[0],-1,self.patchsize**2)
        x2 = x2.transpose(-1, -2)
        wa_L = self.token_wA_L.expand(x1.shape[0],-1,-1)
        wa_L = rearrange(wa_L, 'b h w -> b w h')  # Transpose
        A_L = torch.einsum('bij,bjk->bik', x2, wa_L)
        A_L = rearrange(A_L, 'b h w -> b w h')  # Transpose
        A_L = A_L.softmax(dim=-1)
        wv_L = self.token_wV_L.expand(x2.shape[0],-1,-1)
        VV_L = torch.einsum('bij,bjk->bik', x2, wv_L)
        x2 = torch.einsum('bij,bjk->bik', A_L, VV_L)
        x1 = x1.flatten(2)
        
        x1 = x1.transpose(-1, -2)
        wa = self.token_wA.expand(x1.shape[0],-1,-1)
        wa = rearrange(wa, 'b h w -> b w h')  # Transpose
        A = torch.einsum('bij,bjk->bik', x1, wa)
        A = rearrange(A, 'b h w -> b w h')  # Transpose
        A = A.softmax(dim=-1)
        wv = self.token_wV.expand(x1.shape[0],-1,-1)
        VV = torch.einsum('bij,bjk->bik', x1, wv)
        T = torch.einsum('bij,bjk->bik', A, VV)
        x = torch.cat((x2, T), dim = 1) #[b,n+1,dim]
        embeddings = x + self.position_embeddings
        embeddings = self.dropout(embeddings)
        x = self.ca(embeddings)
        x = x.reshape(x.shape[0],-1)
        out3 = self.out3(x)
        return out3
  




# In[10]:


class Minato(nn.Module):
    def __init__(self,  Classes,c1,c2):
      # x1_shape： B C (P P)
        super(Minato, self).__init__()
        self.HSIOnly = section.getboolean('HSIOnly')
        #临时变量
        self.out_channels = 8
        self.kernel_size = getResolution(section["kernel_size"])
        self.padding_size = getResolution(section["padding_size"])
        self.token_num = int(section['token_num']) # token number / 2
        self.token_dim = int(section['token_dim'])
        self.cnn_out_dim = self.token_dim # origin token dim
        self.center_patch_size = int(section['center_patch_size'])
        self.cnn1d_out_dim = int(section['cnn1d_out_dim'])
        self.cnn1d_kernel1 = int(section['cnn1d_kernel1'])
        self.cnn1d_kernel2 = int(section['cnn1d_kernel2'])
        
        self.patch_size = int(section['patchsize'])
        self.all_token_num = self.getTokenNum()
        self.attn=None
        self.v=None

        self.conv3d_features = nn.Sequential(
            nn.Conv3d(1, self.out_channels, kernel_size=self.kernel_size,padding = self.padding_size,stride = 1),
            nn.BatchNorm3d(self.out_channels),
            nn.ReLU(),
        )
        self.conv2d_features = nn.Sequential(
            nn.Conv2d(in_channels=self.out_channels*self.calDim(inn=c1,k=self.kernel_size[0],s=1,p=self.padding_size[0]), out_channels=self.cnn_out_dim, kernel_size=(3, 3)),
            nn.BatchNorm2d(self.cnn_out_dim),
            nn.ReLU(),
        )
        self.conv2d_features_2 = nn.Sequential(
          nn.Conv2d(in_channels=c2, out_channels=self.cnn_out_dim, kernel_size=(3, 3)),
          nn.BatchNorm2d(self.cnn_out_dim),
          nn.ReLU(),
        )
        
        self.conv1d_features_1 = nn.Sequential(
          nn.Conv1d(in_channels=self.center_patch_size**2, out_channels=self.cnn1d_out_dim, kernel_size=self.cnn1d_kernel1,padding = 1),
          nn.BatchNorm1d(self.cnn1d_out_dim),
          nn.ReLU(),
        )
        
        self.conv1d_features_2 = nn.Sequential(
          nn.Conv1d(in_channels=self.center_patch_size**2, out_channels=self.cnn1d_out_dim, kernel_size=self.cnn1d_kernel2,padding =1),
          nn.BatchNorm1d(self.cnn1d_out_dim),
          nn.ReLU(),
        )
        
        
        self.softmax = nn.Softmax(dim=-1)
        self.wa_1 = nn.Parameter(torch.empty(1, self.token_num, self.cnn_out_dim), requires_grad=True)  # Tokenization parameters
        self.wb_1 = nn.Parameter(torch.empty(1, self.token_dim, self.cnn_out_dim), requires_grad=True)  # Tokenization parameters
        self.wa_2 = nn.Parameter(torch.empty(1, self.token_num, self.cnn_out_dim), requires_grad=True)  # Tokenization parameters
        self.wb_2 = nn.Parameter(torch.empty(1, self.token_dim, self.cnn_out_dim), requires_grad=True)  # Tokenization parameters 
        self.learnableq1 = nn.Parameter(torch.empty(1, 1, self.cnn_out_dim), requires_grad=True)  # Tokenization parameters
        self.learnableq2 = nn.Parameter(torch.empty(1,1, self.cnn_out_dim), requires_grad=True)  # Tokenization parameters
        torch.nn.init.xavier_normal_(self.wa_1)
        torch.nn.init.xavier_normal_(self.wa_2)
        torch.nn.init.xavier_normal_(self.learnableq1)
        torch.nn.init.xavier_normal_(self.learnableq2)

        if section.getboolean('one_wb'):
            torch.nn.init.ones_(self.wb_1)
            torch.nn.init.ones_(self.wb_2)
        else:
            torch.nn.init.xavier_normal_(self.wb_1)
            torch.nn.init.xavier_normal_(self.wb_2)
        # attention
        self.emb_heads=int(section['emb_heads'])
        self.emb_heads_dim=int(section['emb_heads_dim'])
        if section.getboolean('noconv'):
            self.embeddinglinear1 = nn.Linear(c1,self.cnn_out_dim,bias=False)
            self.embeddinglinear2 = nn.Linear(c2,self.cnn_out_dim,bias=False)

        self.qkvlinear1 = nn.Linear(self.cnn_out_dim,3*self.emb_heads*self.emb_heads_dim,bias=False)
        self.qkvlinear2 = nn.Linear(self.cnn_out_dim,3*self.emb_heads*self.emb_heads_dim,bias=False)
        
        self.pxlinear1=nn.Linear(self.cnn1d_out_dim*self.calDim(c1,1,self.cnn1d_kernel1,1),self.token_dim,bias=False)
        self.pxlinear2=nn.Linear(self.cnn1d_out_dim*self.calDim(c2,1,self.cnn1d_kernel2,1),self.token_dim,bias=False)
      
        self.qlinear1=nn.Linear(self.token_dim,self.emb_heads*self.emb_heads_dim,bias=False)
        self.qlinear2=nn.Linear(self.token_dim,self.emb_heads*self.emb_heads_dim,bias=False)

        self.qlinear1_d=nn.Linear(self.cnn_out_dim,self.emb_heads*self.emb_heads_dim,bias=False)
        self.qlinear2_d=nn.Linear(self.cnn_out_dim,self.emb_heads*self.emb_heads_dim,bias=False)

        self.wp1=nn.Linear((self.patch_size-2)**2,1,bias=False)
        self.wp2=nn.Linear((self.patch_size-2)**2,1,bias=False)

        self.vlinear1=nn.Linear(self.cnn_out_dim,self.emb_heads*self.emb_heads_dim,bias=False)
        self.vlinear2=nn.Linear(self.cnn_out_dim,self.emb_heads*self.emb_heads_dim,bias=False)
        self.relinear1 = nn.Linear(self.cnn_out_dim,self.emb_heads*self.emb_heads_dim,bias=False)
        self.relinear2 = nn.Linear(self.cnn_out_dim,self.emb_heads*self.emb_heads_dim,bias=False)
        # self.onelinear1 = nn.Linear(pp,1,bias=False)
        # self.onelinear2 = nn.Linear(pp,1,bias=False)
        
        self.scale = self.emb_heads_dim ** -0.5
        if  section['pos_emb']=='random':
            self.pos = nn.Parameter(torch.randn(1, self.all_token_num+1, self.token_dim))
        if  section['pos_emb']=='concact':
            self.pos = nn.Parameter(torch.randn(1, self.all_token_num+1, self.token_dim))
        if section['cls'] == 'cls':
            self.cls_token = nn.Parameter(torch.randn(1, 1, self.token_dim))
        self.dropout = nn.Dropout(0.1)
        encoder_layer = nn.TransformerEncoderLayer(d_model=self.token_dim, nhead=8,batch_first=True,dim_feedforward=int(section['dim_feedforward']))
        self.transformer =  nn.TransformerEncoder(encoder_layer, num_layers=int(section['transformer_layer_num']),)
        self.norm = nn.LayerNorm(self.token_dim)
        self.gap = nn.AdaptiveAvgPool1d(1)
        self.mlp_head = nn.Sequential(
            nn.LayerNorm(self.token_dim),
            nn.Linear(self.token_dim, Classes)
        )
        
        
    def forward(self, x1, x2):
        p = self.patch_size
        cp = self.center_patch_size
        
        b, c, n = x1.shape
        # print(x1.shape,x2.shape)
        x1 = rearrange(x1, 'b c (h w) -> b c h w',h=p)
        x2 = rearrange(x2, 'b c (h w) -> b c h w',h=p)
        
        x1p = x1[:,:,p//2-cp//2:p//2+cp//2+1,p//2-cp//2:p//2+cp//2+1]
        x2p = x2[:,:,p//2-cp//2:p//2+cp//2+1,p//2-cp//2:p//2+cp//2+1]
        x1p = rearrange(x1p, 'b c h w -> b (h w) c') # 1 c1
        x2p = rearrange(x2p, 'b c h w -> b (h w) c') # 1 c2
        if section.getboolean('noconv'):
            x1 = rearrange(x1, 'b c h w -> b (h w) c')
            
            x1 = self.embeddinglinear1(x1)
            
            x2 = rearrange(x2, 'b c h w -> b (h w) c')
            x2 = self.embeddinglinear2(x2)
        else:#conv
            # x1 -> conv3d -> conv2d
            x1 = rearrange(x1, 'b c h w -> b 1 c h w',h=p)
            x1 = self.conv3d_features(x1)
            x1 = rearrange(x1, 'b c d h w -> b (c d) h w')
            x1 = self.conv2d_features(x1)
            x1 = rearrange(x1,'b c h w -> b (h w) c') #x1 tokens | B x1_shape[2] cnn_out_dim

            # x2 -> conv2d
            x2 = self.conv2d_features_2(x2)
            x2 = rearrange(x2, 'b c h w -> b (h w) c')#x2 tokens | B x2_shape[2] cnn_out_dim
        if section['to'] == 'scfem':
            x = torch.concat((x1,x2), dim=1)
            x = self.qlinear1_d(x)
            x = self.tokenToCls(x)
            return x

        if section['cls'] == 'fourgap':
            x1 = rearrange(x1,'b h w -> b w h')# B dim tokennumber
            x1 = self.gap(x1)
            x1 = rearrange(x1,'b h w -> b w h')# B 1 dim
            x2 = rearrange(x2,'b h w -> b w h')# B dim tokennumber
            x2 = self.gap(x2)
            x2 = rearrange(x2,'b h w -> b w h')# B 1 dim
        if section.getboolean('nosoftmax'):
            x = torch.concat((x1,x2), dim=1)
        else:
            x = self.getTokens(x1,x2,x1p,x2p)
        if section['to'] == 'cpscfem' or  section['to'] == '1drcatm' :
            x = self.qlinear1_d(x)
            x = self.tokenToCls(x)
            return x
        if section['cls'] == 'cls':
            cls_tokens = repeat(self.cls_token, '() n d -> b n d', b = x.shape[0]) #[b,1,dim]
            x = torch.concat((cls_tokens,x), dim=1)
        
            
        self.input_tokens = x[0].unsqueeze(0)
        
        if  section['pos_emb']!='none':
            x = x+self.pos[:,:x.shape[1],:]
        
        x = self.dropout(x)
        if not section.getboolean('notrans'):
            x = self.transformer(x)
        self.output_tokens = x[0].unsqueeze(0)
        
        x = self.norm(x)
        self.normed_tokens = x[0].unsqueeze(0)
        
        if section['cls'] == 'cls':
            x1,x2,fusion,_ = self.getSplitToken(x[:,1:,:])
            x= x[:,0,:].squeeze(1)
            l = [x1,x2,fusion]
            for i in range(len(l)):
                l[i] = self.tokenToCls(l[i])
            x1,x2,fusion = l
            
            return x1,x2,fusion,x
        
        elif section['cls'] == 'gap':
            
            
            x1,x2,fusion,x = self.getSplitToken(x)
            l = [x1,x2,fusion,x]
            for i in range(len(l)):
                l[i] = self.tokenToCls(l[i])
            x1,x2,fusion,x = l

            return x1,x2,fusion,x
        elif section['cls'] == 'fourgap':
            x = self.tokenToCls(x)
            return None,None,None,x
        elif section['cls'] == 'gapx':
            x = self.tokenToCls(x)

            return x
    
    
    def calDim(self,inn,p,k,s):
        return int((inn+2*p-(k-1)-1)/s+1)
    
    def linear_attn(self,q,k,v,):
        # b d n
        attended_values = []
        attended_value = None
        for i in range(self.emb_heads):
            kk = F.softmax(k[
                :,
                i * self.emb_heads_dim: (i + 1) * self.emb_heads_dim,
                :
            ], dim=2)
            qq = F.softmax(q[
                :,
                i * self.emb_heads_dim: (i + 1) * self.emb_heads_dim,
                :
            ], dim=1)
            vv = v[
                :,
                i * self.emb_heads_dim: (i + 1) * self.emb_heads_dim,
                :
            ]
            context = kk @ vv.transpose(1, 2) # d * d
            attended_value = (
                context.transpose(1, 2) @ qq # b * d * n  
            )
            attended_values.append(attended_value)
        aggregated_values = torch.cat(attended_values, dim=1)
        aggregated_values = rearrange(aggregated_values,'b d n -> b n d')
        return aggregated_values
    
    def getTokens(self,x1,x2,x1p,x2p):
        select_token_mode = section['select_token_mode']
        self.attn=[]
        self.v=[]
        if section['to'] == 'cpscfem':
            q1 = self.conv1d_features_1(x1p) # 1 c ->(1@3 p=1 s=1) -> 1 c1
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')# 1 c1
            q1 = self.qlinear1(q1) # b 1 dim
            q2 = self.conv1d_features_2(x2p) # 1 c ->(1@3 p=1 s=1) -> 1 c1
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')# 1 c1
            q2 = self.qlinear2(q2) # b 1 dim
            x = torch.concat((x1,x2,q1,q2), dim=1)
            return x
        if select_token_mode == '11,22,12,21':
            # w^T
            wa_1 = rearrange(self.wa_1,'b h w -> b w h')# 1,  CNNdim, N
            wb_1 = rearrange(self.wb_1,'b h w -> b w h')# 1,  CNNdim, dim
            wa_2 = rearrange(self.wa_2,'b h w -> b w h')# 1,  CNNdim, N
            wb_2 = rearrange(self.wb_2,'b h w -> b w h')# 1,  CNNdim, dim

            #softmax(x1*wa1)T
            tmp1 = torch.einsum('bij,bjk->bik', x1, wa_1)# x1*wa1           | B x2_shape[2]    N
            tmp1 = self.softmax(tmp1)                    # softmax(x1*wa1)  | B x2_shape[2]  dim
            tmp1 = rearrange(tmp1,'b h w -> b w h')      # softmax(x1*wa1)T | B N    x2_shape[2]

            #x1wb1
            tmp2 = torch.einsum('bij,bjk->bik', x1, wb_1) # x1*wb1          | B x2_shape[2] dim

            #softmax(x2*wa2)T
            tmp3 = torch.einsum('bij,bjk->bik', x2, wa_2)
            tmp3 = self.softmax(tmp3)  
            tmp3 = rearrange(tmp3,'b h w -> b w h')

            #x2wb2
            tmp4 = torch.einsum('bij,bjk->bik', x2, wb_2)


            #softmax(x1*wa1)T * x1wb1 [ 1 * 1 ] 
            x1 = torch.einsum('bij,bjk->bik', tmp1, tmp2) # B N dim
            #softmax(x2*wa2)T * x2wb2 [ 2 * 2 ] 
            x2 = torch.einsum('bij,bjk->bik', tmp3, tmp4) # B N dim
            #softmax(x1*wa1)T * x2wb2 [ 1 * 2 ] 
            tmp1 = torch.einsum('bij,bjk->bik', tmp1, tmp4) # B N dim
            #softmax(x2*wa2)T * x1wb1 [ 2 * 1 ] 
            tmp2 = torch.einsum('bij,bjk->bik', tmp3, tmp2) # B N dim
            
            x = torch.concat((x1,x2,tmp1,tmp2), dim=1)
            
            return x
        if select_token_mode == '1,2,12,21':
            # w^T
            wa_1 = rearrange(self.wa_1,'b h w -> b w h')# 1,  CNNdim, N
            wb_1 = rearrange(self.wb_1,'b h w -> b w h')# 1,  CNNdim, dim
            wa_2 = rearrange(self.wa_2,'b h w -> b w h')# 1,  CNNdim, N
            wb_2 = rearrange(self.wb_2,'b h w -> b w h')# 1,  CNNdim, dim

            #softmax(x1*wa1)T
            tmp1 = torch.einsum('bij,bjk->bik', x1, wa_1)# x1*wa1           | B x2_shape[2]    N
            tmp1 = self.softmax(tmp1)                    # softmax(x1*wa1)  | B x2_shape[2]  dim
            tmp1 = rearrange(tmp1,'b h w -> b w h')      # softmax(x1*wa1)T | B N    x2_shape[2]

            #x1wb1
            tmp2 = torch.einsum('bij,bjk->bik', x1, wb_1) # x1*wb1          | B x2_shape[2] dim

            #softmax(x2*wa2)T
            tmp3 = torch.einsum('bij,bjk->bik', x2, wa_2)
            tmp3 = self.softmax(tmp3)  
            tmp3 = rearrange(tmp3,'b h w -> b w h')

            #x2wb2
            tmp4 = torch.einsum('bij,bjk->bik', x2, wb_2)

            #softmax(x1*wa1)T * x2wb2 [ 1 * 2 ] 
            tmp1 = torch.einsum('bij,bjk->bik', tmp1, tmp4) # B N dim
            #softmax(x2*wa2)T * x1wb1 [ 2 * 1 ] 
            tmp2 = torch.einsum('bij,bjk->bik', tmp3, tmp2) # B N dim
            
            x = torch.concat((x1,x2,tmp1,tmp2), dim=1)
            
            return x
        if select_token_mode == '1,2':
            x = torch.concat((x1,x2), dim=1)
            return x
        if select_token_mode == '12,21':
            # w^T
            wa_1 = rearrange(self.wa_1,'b h w -> b w h')# 1,  CNNdim, N
            wb_1 = rearrange(self.wb_1,'b h w -> b w h')# 1,  CNNdim, dim
            wa_2 = rearrange(self.wa_2,'b h w -> b w h')# 1,  CNNdim, N
            wb_2 = rearrange(self.wb_2,'b h w -> b w h')# 1,  CNNdim, dim

            #softmax(x1*wa1)T
            tmp1 = torch.einsum('bij,bjk->bik', x1, wa_1)# x1*wa1           | B x2_shape[2]    N
            tmp1 = self.softmax(tmp1)                    # softmax(x1*wa1)  | B x2_shape[2]  dim
            tmp1 = rearrange(tmp1,'b h w -> b w h')      # softmax(x1*wa1)T | B N    x2_shape[2]

            #x1wb1
            tmp2 = torch.einsum('bij,bjk->bik', x1, wb_1) # x1*wb1          | B x2_shape[2] dim

            #softmax(x2*wa2)T
            tmp3 = torch.einsum('bij,bjk->bik', x2, wa_2)
            tmp3 = self.softmax(tmp3)  
            tmp3 = rearrange(tmp3,'b h w -> b w h')

            #x2wb2
            tmp4 = torch.einsum('bij,bjk->bik', x2, wb_2)

            #softmax(x1*wa1)T * x2wb2 [ 1 * 2 ] 
            tmp1 = torch.einsum('bij,bjk->bik', tmp1, tmp4) # B N dim
            #softmax(x2*wa2)T * x1wb1 [ 2 * 1 ] 
            tmp2 = torch.einsum('bij,bjk->bik', tmp3, tmp2) # B N dim
            
            x = torch.concat((tmp1,tmp2), dim=1)
            
            return x
        if select_token_mode == '11,22':
            # w^T
            wa_1 = rearrange(self.wa_1,'b h w -> b w h')# 1,  CNNdim, N
            wb_1 = rearrange(self.wb_1,'b h w -> b w h')# 1,  CNNdim, dim
            wa_2 = rearrange(self.wa_2,'b h w -> b w h')# 1,  CNNdim, N
            wb_2 = rearrange(self.wb_2,'b h w -> b w h')# 1,  CNNdim, dim

            #softmax(x1*wa1)T
            tmp1 = torch.einsum('bij,bjk->bik', x1, wa_1)# x1*wa1           | B x2_shape[2]    N
            tmp1 = self.softmax(tmp1)                    # softmax(x1*wa1)  | B x2_shape[2]  dim
            tmp1 = rearrange(tmp1,'b h w -> b w h')      # softmax(x1*wa1)T | B N    x2_shape[2]

            #x1wb1
            tmp2 = torch.einsum('bij,bjk->bik', x1, wb_1) # x1*wb1          | B x2_shape[2] dim

            #softmax(x2*wa2)T
            tmp3 = torch.einsum('bij,bjk->bik', x2, wa_2)
            tmp3 = self.softmax(tmp3)  
            tmp3 = rearrange(tmp3,'b h w -> b w h')

            #x2wb2
            tmp4 = torch.einsum('bij,bjk->bik', x2, wb_2)

            #softmax(x1*wa1)T * x1wb1 [ 1 * 1 ] 
            x1 = torch.einsum('bij,bjk->bik', tmp1, tmp2) # B N dim
            #softmax(x2*wa2)T * x2wb2 [ 2 * 2 ] 
            x2 = torch.einsum('bij,bjk->bik', tmp3, tmp4) # B N dim
            
            x = torch.concat((x1,x2), dim=1)
            
            return x
        if select_token_mode == 'q2k1v1,q1k2v2':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q2k1v1
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q1k2v2
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            x = torch.concat((q2k1v1,q1k2v2), dim=1)
            
            return x
        if select_token_mode == 'q2k1v1':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q2k1v1
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            return q2k1v1

        if select_token_mode == 'q1k2v2':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k2v2
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            return q1k2v2

        if select_token_mode == 'q1k1v1,q2k1v1':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
            
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            # q2k1v1
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)

            x = torch.concat((q1k1v1,q2k1v1), dim=1)
            
            return x
        if select_token_mode == 'q1k1v1,q2k2v2':
            
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            x = torch.concat((q1k1v1,q2k2v2), dim=1)
            
            return x
        if select_token_mode == 'q1k1v1,q2k2v2,q2k1v1,q1k2v2':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'

            
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            
            return x
        if select_token_mode == 'v1,v2,v1,v2':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'

            v1=self.vlinear1(x1)
            v1=rearrange(v1, 'b n (h d) -> b h n d', h=self.emb_heads)
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            
            return x
        if select_token_mode == 'spectralq':
            q1 = self.conv1d_features_1(x1p) # b (h w) c -> 
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)
            
            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)
            
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            return x
        if select_token_mode == 'spectralq_res':
            res1 = self.relinear1(x1)
            res2 = self.relinear2(x2)
            q1 = self.conv1d_features_1(x1p)
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)
            
            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)
            
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            return x
        
        
        if select_token_mode == 'linear_attn_spectralq':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
        
            q1 = self.conv1d_features_1(x1p)
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n d -> b d n')
            
            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2, 'b n d -> b d n')
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n d -> b d n'), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n d -> b d n'), qkv2)
            
            self.v.append(q1[0])
            self.v.append(q2[0])
            self.v.append(k1[0])
            self.v.append(k2[0])
            self.v.append(v1[0])
            self.v.append(v2[0])
            
            q1k1v1 = self.linear_attn(q1,k1,v1)
            q2k2v2 = self.linear_attn(q2,k2,v2)
            q2k1v1 = self.linear_attn(q2,k1,v1)
            q1k2v2 = self.linear_attn(q1,k2,v2)
            

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            
            return x
    
        if select_token_mode == 'linear_attn_spectralq_res':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
            res1 = self.relinear1(x1)
            res2 = self.relinear2(x2)
            q1 = self.conv1d_features_1(x1p)
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n d -> b d n')
            
            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2, 'b n d -> b d n')
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n d -> b d n'), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n d -> b d n'), qkv2)
            
            self.v.append(q1[0])
            self.v.append(q2[0])
            self.v.append(k1[0])
            self.v.append(k2[0])
            self.v.append(v1[0])
            self.v.append(v2[0])
            
            q1k1v1 = self.linear_attn(q1,k1,v1) + res1
            q2k2v2 = self.linear_attn(q2,k2,v2) + res2
            q2k1v1 = self.linear_attn(q2,k1,v1) + res1
            q1k2v2 = self.linear_attn(q1,k2,v2) + res2
            

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            
            return x
        if select_token_mode == 'linear_attn':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
        
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n d -> b d n'), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n d -> b d n'), qkv2)
            
            self.v.append(q1[0])
            self.v.append(q2[0])
            self.v.append(k1[0])
            self.v.append(k2[0])
            self.v.append(v1[0])
            self.v.append(v2[0])
            # q1k1v1
            q1k1v1 = self.linear_attn(q1,k1,v1)
            q2k2v2 = self.linear_attn(q2,k2,v2)
            q2k1v1 = self.linear_attn(q2,k1,v1)
            q1k2v2 = self.linear_attn(q1,k2,v2)
            

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            
            return x
        if select_token_mode == 'linear_attn_res':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
        
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n d -> b d n'), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n d -> b d n'), qkv2)
            res1 = self.relinear1(x1)
            res2 = self.relinear2(x2)
            self.v.append(q1[0])
            self.v.append(q2[0])
            self.v.append(k1[0])
            self.v.append(k2[0])
            self.v.append(v1[0])
            self.v.append(v2[0])
            # q1k1v1
            q1k1v1 = self.linear_attn(q1,k1,v1)+res1
            q2k2v2 = self.linear_attn(q2,k2,v2)+res2
            q2k1v1 = self.linear_attn(q2,k1,v1)+res1
            q1k2v2 = self.linear_attn(q1,k2,v2)+res2
            

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            
            return x
        if select_token_mode == 'q1k1v1,q2k2v2,q2k1v1,q1k2v2_res':
            assert x1.shape[1]==x2.shape[1] , f'x1.shape[1]]={x1.shape[1]},x2.shape[1]={x2.shape[1]} x1!=x2'
        
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            res1 = self.relinear1(x1)
            res2 = self.relinear2(x2)
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads) + res1
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads) + res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads) + res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads) + res2

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            
            return x
        if select_token_mode == 'v1v2':
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, _, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, _, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            return x
        if select_token_mode == 'spectralq_res_one':
            x1 = rearrange(x1, 'b n c -> b c n')
            res1 = self.gap(x1)
            x1 = rearrange(x1, 'b c n -> b n c')
            
            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            x2 = rearrange(x2, 'b c n -> b n c')
            
            q1 = self.conv1d_features_1(x1p)
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)
            
            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)
            
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            # print(x.shape)
            return x
        if select_token_mode == 'spectralq_res_one_true':
            x1 = rearrange(x1, 'b n c -> b c n') #  d pp
            res1 = self.gap(x1) # d 1
            res1 = rearrange(res1, 'b n c -> b c n') # 1 d
            # print("res1 shape:",res1.shape)
            x1 = rearrange(x1, 'b c n -> b n c') # pp d
            
            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            res2 = rearrange(res2, 'b n c -> b c n')
            # print("res2 shape:",res2.shape)
            x2 = rearrange(x2, 'b c n -> b n c')
            
            q1 = self.conv1d_features_1(x1p) # 1 c ->(1@3 p=1 s=1) -> 1 c1
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')# 1 c1
            q1 = self.pxlinear1(q1)
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)
            
            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.pxlinear2(q2)
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)
            
            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)
            
            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1
            
            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            # print(x.shape)
            return x
        if select_token_mode == 'cpscfem-gap':
            x1 = rearrange(x1, 'b n c -> b c n') #  d pp
            res1 = self.gap(x1) # d 1
            res1 = rearrange(res1, 'b n c -> b c n') # 1 d
            # print("res1 shape:",res1.shape)
            x1 = rearrange(x1, 'b c n -> b n c') # pp d

            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            res2 = rearrange(res2, 'b n c -> b c n')
            # print("res2 shape:",res2.shape)
            x2 = rearrange(x2, 'b c n -> b n c')

            q1 = self.qlinear1_d(res1)
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)

            q2 = self.qlinear2_d(res2)
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)

            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)

            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            # print(x.shape)
            return x
        if select_token_mode == 'cpscfem-wp':
            x1 = rearrange(x1, 'b n c -> b c n') #  d pp
            res1 = self.gap(x1) # d 1
            qq1 = self.wp1(x1) # d 1
            qq1 = rearrange(qq1, 'b n c -> b c n') # 1 d
            res1 = rearrange(res1, 'b n c -> b c n') # 1 d
            # print("res1 shape:",res1.shape)
            x1 = rearrange(x1, 'b c n -> b n c') # pp d

            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            qq2 = self.wp2(x2) # d 1
            qq2 = rearrange(qq2, 'b n c -> b c n') # 1 d
            res2 = rearrange(res2, 'b n c -> b c n')
            # print("res2 shape:",res2.shape)
            x2 = rearrange(x2, 'b c n -> b n c')

            q1 = self.qlinear1_d(qq1)
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)

            q2 = self.qlinear2_d(qq2)
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)

            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)

            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            # print(x.shape)
            return x
        if select_token_mode == 'cpscfem-learnable':
            x1 = rearrange(x1, 'b n c -> b c n') #  d pp
            res1 = self.gap(x1) # d 1
            res1 = rearrange(res1, 'b n c -> b c n') # 1 d
            # print("res1 shape:",res1.shape)
            x1 = rearrange(x1, 'b c n -> b n c') # pp d

            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            res2 = rearrange(res2, 'b n c -> b c n')
            # print("res2 shape:",res2.shape)
            x2 = rearrange(x2, 'b c n -> b n c')

            q1 = self.qlinear1_d(self.learnableq1)
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)

            q2 = self.qlinear2_d(self.learnableq2)
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)

            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)

            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            # print(x.shape)
            return x
        if select_token_mode == '1drca-nores':
            x1 = rearrange(x1, 'b n c -> b c n') #  d pp
            res1 = self.gap(x1) # d 1
            res1 = rearrange(res1, 'b n c -> b c n') # 1 d
            # print("res1 shape:",res1.shape)
            x1 = rearrange(x1, 'b c n -> b n c') # pp d

            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            res2 = rearrange(res2, 'b n c -> b c n')
            # print("res2 shape:",res2.shape)
            x2 = rearrange(x2, 'b c n -> b n c')

            q1 = self.conv1d_features_1(x1p) # 1 c ->(1@3 p=1 s=1) -> 1 c1
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')# 1 c1
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)

            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)

            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)

            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)

            x = torch.concat((q1k1v1,q2k2v2,q2k1v1,q1k2v2), dim=1)
            # print(x.shape)
            return x
        if select_token_mode == '1drca-justmsa':
            x1 = rearrange(x1, 'b n c -> b c n') #  d pp
            res1 = self.gap(x1) # d 1
            res1 = rearrange(res1, 'b n c -> b c n') # 1 d
            # print("res1 shape:",res1.shape)
            x1 = rearrange(x1, 'b c n -> b n c') # pp d

            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            res2 = rearrange(res2, 'b n c -> b c n')
            # print("res2 shape:",res2.shape)
            x2 = rearrange(x2, 'b c n -> b n c')

            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            q1, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            q2, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)

            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k1v1,q2k2v2), dim=1)

            return x
        if select_token_mode == '1drca-1122':
            x1 = rearrange(x1, 'b n c -> b c n') #  d pp
            res1 = self.gap(x1) # d 1
            res1 = rearrange(res1, 'b n c -> b c n') # 1 d
            # print("res1 shape:",res1.shape)
            x1 = rearrange(x1, 'b c n -> b n c') # pp d

            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            res2 = rearrange(res2, 'b n c -> b c n')
            # print("res2 shape:",res2.shape)
            x2 = rearrange(x2, 'b c n -> b n c')

            q1 = self.conv1d_features_1(x1p) # 1 c ->(1@3 p=1 s=1) -> 1 c1
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')# 1 c1
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)

            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)

            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)

            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k1v1,q2k2v2), dim=1)
            # print(x.shape)
            return x
        if select_token_mode == '1drca-1221':
            x1 = rearrange(x1, 'b n c -> b c n') #  d pp
            res1 = self.gap(x1) # d 1
            res1 = rearrange(res1, 'b n c -> b c n') # 1 d
            # print("res1 shape:",res1.shape)
            x1 = rearrange(x1, 'b c n -> b n c') # pp d

            x2 = rearrange(x2, 'b n c -> b c n')
            res2 = self.gap(x2)
            res2 = rearrange(res2, 'b n c -> b c n')
            # print("res2 shape:",res2.shape)
            x2 = rearrange(x2, 'b c n -> b n c')

            q1 = self.conv1d_features_1(x1p) # 1 c ->(1@3 p=1 s=1) -> 1 c1
            q1 = rearrange(q1, 'b n c -> b 1 (n c)')# 1 c1
            q1 = self.qlinear1(q1) # b 1 dim
            q1=rearrange(q1, 'b n (h d) -> b h n d', h=self.emb_heads)

            q2 = self.conv1d_features_2(x2p)
            q2 = rearrange(q2, 'b n c -> b 1 (n c)')
            q2 = self.qlinear2(q2) # b 1 dim
            q2=rearrange(q2,'b n (h d) -> b h n d', h=self.emb_heads)

            qkv1=self.qkvlinear1(x1).chunk(3, dim=-1)
            _, k1, v1 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv1)
            qkv2=self.qkvlinear2(x2).chunk(3, dim=-1)
            _, k2, v2 = map(lambda t: rearrange(t, 'b n (h d) -> b h n d', h=self.emb_heads), qkv2)

            # q1k1v1
            q1 = q1*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q1, k1)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q1k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q1k1v1 = rearrange(q1k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q2k2v2
            q2 = q2*self.scale
            dots = torch.einsum('bhid,bhjd->bhij', q2, k2)
            attn = dots.softmax(dim=-1)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q2k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q2k2v2 = rearrange(q2k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            # q2k1v1
            dots = torch.einsum('bhid,bhjd->bhij', q2, k1)
            self.attn.append(attn[0])
            self.v.append(v1[0])
            q2k1v1 = torch.einsum('bhij,bhjd->bhid', attn, v1)
            q2k1v1 = rearrange(q2k1v1, 'b h n d -> b n (h d)', h=self.emb_heads)+res1

            # q1k2v2
            dots = torch.einsum('bhid,bhjd->bhij', q1, k2)
            self.attn.append(attn[0])
            self.v.append(v2[0])
            q1k2v2 = torch.einsum('bhij,bhjd->bhid', attn, v2)
            q1k2v2 = rearrange(q1k2v2, 'b h n d -> b n (h d)', h=self.emb_heads)+res2

            x = torch.concat((q1k2v2,q2k1v1), dim=1)
            # print(x.shape)
            return x
    def getTokenNum(self):
        select_token_mode = section['select_token_mode']
        if select_token_mode == '11,22,12,21':
            return self.token_num*4
        if select_token_mode == '1,2,12,21':
            ps=int(self.patch_size)
            self.x1_n = self.calDim(inn=ps,k=self.kernel_size[1],p=self.padding_size[1],s=1)
            self.x1_n = self.calDim(inn=self.x1_n,k=3,p=0,s=1)**2
            self.x2_n = self.calDim(inn=ps,k=3,p=0,s=1)**2
            return self.x1_n+self.x2_n+2*self.token_num
        if select_token_mode == '1,2':
            ps=int(self.patch_size)
            self.x1_n = self.calDim(inn=ps,k=self.kernel_size[1],p=self.padding_size[1],s=1)
            self.x1_n = self.calDim(inn=self.x1_n,k=3,p=0,s=1)**2
            self.x2_n = self.calDim(inn=ps,k=3,p=0,s=1)**2
            return self.x1_n+self.x2_n
        if select_token_mode == '12,21':
            return 2*self.token_num
        if select_token_mode == '11,22':
            return 2*self.token_num
        if select_token_mode == 'q1k1v1,q2k1v1':
            ps=int(self.patch_size)
            self.x1_n = self.calDim(inn=ps,k=self.kernel_size[1],p=self.padding_size[1],s=1)
            self.x1_n = self.calDim(inn=self.x1_n,k=3,p=0,s=1)**2
            self.x2_n = self.calDim(inn=ps,k=3,p=0,s=1)**2
            return self.x1_n+self.x2_n
        if select_token_mode[0]=='q':
            ps=int(self.patch_size)
            l = select_token_mode.split(',')
            self.x1_n = self.calDim(inn=ps,k=self.kernel_size[1],p=self.padding_size[1],s=1)
            self.x1_n = self.calDim(inn=self.x1_n,k=3,p=0,s=1)**2
            self.x2_n = self.calDim(inn=ps,k=3,p=0,s=1)**2
            assert self.x1_n==self.x2_n,"self.x1_n!=self.x2_n"
            return len(l)*self.x2_n
        if select_token_mode=='spectralq_res_one_true':
            return 4
            
    def getSplitToken(self,x):
        select_token_mode = section['select_token_mode']
        if select_token_mode == '11,22,12,21':
            return x[:,0:self.token_num,:],x[:,self.token_num:self.token_num*2,:],x[:,self.token_num*2:,:],x
        if select_token_mode == '1,2,12,21':
            return x[:,0:self.x1_n,:],x[:,self.x1_n:self.x1_n+self.x2_n,:],x[:,self.x1_n+self.x2_n:,:],x
        if select_token_mode == '1,2':
            return x[:,0:self.x1_n,:],x[:,self.x1_n:self.x1_n+self.x2_n,:],None,x
        if select_token_mode == '12,21':
            return None,None,x,x
        if select_token_mode == '11,22':
            return x[:,0:self.token_num,:],x[:,self.token_num:self.token_num*2,:],None,x
        else:
            return None,None,None,x
        
    def tokenToCls(self,x):
        if x==None:
            return None
        # x : B tokennumber dim 
        x = rearrange(x,'b h w -> b w h')# B dim tokennumber
        x = self.gap(x)
        x = rearrange(x,'b h w -> b (h w)') # B dim
        x = self.mlp_head(x)
        return x




# In[11]:


# othermethod
from models.exvit import MViT as ExViT
from models.HCTnet import HCTnet as HCT
from models.TDCNN import TDCNN as TDCNN
from models.TD1CNN import TD1CNN as TD1CNN
from models.TD2CNN import TD2CNN as TD2CNN
from models.M2FNet import M2Fnet as M2Fnet

def getModel(name,c1,c2,num_classes,patch_size):
    if name=='ExViT':
        return ExViT(
        patch_size = patch_size,
        num_patches = [c1,c2],
        num_classes = num_classes,
        dim = 64,
        depth = 6,
        heads = 4,
        mlp_dim = 16,
        dropout = 0.1,
        emb_dropout = 0.1,
        mode = 'MViT'
    )
    elif name=='Minato':
        return Minato(c1=c1,
        c2=c2,
        Classes=num_classes
    )
    elif name == 'MFT':
        return MFT(16, c1,c2, num_classes, False,patch_size)
    elif name == 'HCT':
        return HCT(c1=c1,
                   c2=c2,
                   num_classes=num_classes,
                   )
    elif name == '3DCNN':
        return TDCNN(c1=c1,c2=c2,classes=num_classes)
    elif name == '1DCNN':
        return TD1CNN(c1=c1,c2=c2,classes=num_classes,patchsize=patch_size)
    elif name == '2DCNN':
        return TD2CNN(c1=c1,c2=c2,classes=num_classes,patchsize=patch_size)
    elif name == 'M2Fnet':
        return M2Fnet(FM=16, c1=c1, Classes=num_classes,c2=c2,patch_size=patch_size)


# # Train
# 

# In[12]:


#tool
def valid(testLoader,model,is_all=True):
    tmp = None
    pre = np.array([])
    tar = np.array([])
    for step, (b_x1, b_x2, b_y) in enumerate(testLoader):
        # move train data to GPU
        b_x1 = b_x1.cuda()
        
        if section['cls'] == 'gapx':
            if HSIOnly:
                    x = model(b_x1,  b_x2)
                    p = torch.max(x, 1)[1].squeeze()

            else:
                    b_x2 = b_x2.cuda()
                    x= model(b_x1, b_x2)
                    p = torch.max(x, 1)[1].squeeze()
            
        
        else:
            if HSIOnly:
                    x1,fusion,x = model(b_x1,  b_x2)
                    p = torch.max(x, 1)[1].squeeze()

            else:
                    b_x2 = b_x2.cuda()
                    x1,x2,fusion,x= model(b_x1, b_x2)
                    p = torch.max(x, 1)[1].squeeze()

        tar = np.append(tar, b_y.data.numpy()) # 不断append正确标签集合
        pre = np.append(pre, p.data.cpu().numpy()) # 不断append预测标签集合
        if not is_all:
            if tmp == None:
                tmp = 5
            tmp-=1
            if tmp == 0:
                break
    return tar,pre


def cal_loss(x1,x2,fusion,x,loss_func,y):
    namda = float(section['namda'])
    if section['loss_mode']=='x':
        return loss_func(x,y)
    elif section['loss_mode']=='namda(x1,x2,fusion),x':
        return namda*(loss_func(x1,y)+loss_func(x2,y)+loss_func(fusion,y))+loss_func(x,y)
        
print("----------------------------------Training for ",datasetName," ---------------------------------------------")
def show_image(name,step,data):
    w_grid = torchvision.utils.make_grid(data,normalize=True,scale_each=True,padding=2,nrow=int(section['emb_heads']))
    writer.add_image(name, w_grid, global_step=step,dataformats ="CHW")
def train():
    datasetConfig = DatasetConfig(datasetName)
    trainLoader = datasetConfig.getTrainLoader(type = "Tr",batchsize = batchsize)
    testLoader = datasetConfig.getTrainLoader(type = "Te",batchsize = testSizeNumber)

    bestmodel = None

    print("Number of Classes = ", datasetConfig.classNum)
    print("Number of band = ", datasetConfig.bandNum1)
    KAPPA = []
    OA = []
    AA = []
    ELEMENT_ACC = np.zeros((1, datasetConfig.classNum))
    fileName = None

    set_seed(int(section['seed']))
    # summary(model, [(datasetConfig.bandNum1, patchsize**2),(datasetConfig.bandNum2,patchsize**2)])
    for iterNum in range(1):
        print(datasetConfig.bandNum1, datasetConfig.bandNum2, datasetConfig.classNum)
        if section['load_model']!='none':
            print("load file:"+str(checkpointDatasetPath / section['load_model']))
            model = torch.load(checkpointDatasetPath / section['load_model'])
        else:
            model = getModel(section['network'],datasetConfig.bandNum1,datasetConfig.bandNum2,datasetConfig.classNum,int(section['patchsize'])).cuda()
        optimizer = torch.optim.Adam(model.parameters(), lr=LR,weight_decay=5e-3)
        loss_func = nn.CrossEntropyLoss()  # the target label is not one-hotted
        scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=50, gamma=0.9)
        BestAcc = 0

        torch.cuda.synchronize()
        start = time.time()
        # train and test the designed model
        for epoch in range(EPOCH):
                # 测试模型耗时则取消注释
                # prof = torch.profiler.profile(
                #   schedule=torch.profiler.schedule(wait=1, warmup=1, active=3, repeat=1),
                #   on_trace_ready=torch.profiler.tensorboard_trace_handler(tensorboardPath),
                #   record_shapes=True,
                #   with_stack=True)
                # prof.start()
                for step, (b_x1, b_x2, b_y) in enumerate(trainLoader):
                        # 测试模型耗时则取消注释
                        # prof.step()
                        # if step >= 1 + 1 + 3:
                        #     print("return")
                        #     prof.stop() 
                        #     return
                        # move train data to GPU
                        b_x1 = b_x1.cuda()
                        b_y = b_y.cuda()
                        if section['cls'] == 'gapx':
                            if HSIOnly:
                                    x = model(b_x1,  b_x2)
                                    loss = loss_func(x,b_y)
                                    del b_x1,b_y

                            else:
                                    b_x2 = b_x2.cuda()
                                    x= model(b_x1, b_x2)
                                    loss = loss_func(x,b_y)
                                    del b_x1,b_y,b_x2
                        else:
                            if HSIOnly:
                                    x1,x2,fusion,x = model(b_x1,  b_x2)
                                    loss = cal_loss(x1=x1,x2=None,fusion=fusion,x=x,loss_func=loss_func,y = b_y)
                                    del b_x1,b_y

                            else:
                                    b_x2 = b_x2.cuda()
                                    x1,x2,fusion,x= model(b_x1, b_x2)
                                    loss = cal_loss(x1=x1,x2=x2,fusion=fusion,x=x,loss_func=loss_func,y = b_y)
                                    del b_x1,b_y,b_x2
                        writer.add_scalar(f"Loss/train/{iterNum}", loss, epoch*len(trainLoader)+step)   
                        optimizer.zero_grad()  # clear gradients for this training step
                        loss.backward()  # backpropagation, compute gradients
                        if section.getboolean('loss_clip'):
                            nn.utils.clip_grad_norm_(model.parameters(), max_norm=20, norm_type=2)
                        optimizer.step()  # apply gradients



                #每个epoch测一下

                with torch.no_grad():
                    model.eval()
                    tar,pre = valid(testLoader,model)
                    # accuracy = np.sum(pre==tar) / tar.shape[0]

                    oa = accuracy_score(tar, pre)
                    confusion = confusion_matrix(tar, pre,labels=range(datasetConfig.classNum))
                    # print(confusion)
                    each_acc, aa = AA_andEachClassAccuracy(confusion)
                    kappa = cohen_kappa_score(tar, pre)


                    print(np.sum(pre==tar),tar.shape[0])
                    print('Epoch: ', epoch, '| train loss: %.4f' % loss.data.cpu().numpy(), '| test accuracy: %.4f' % (oa*100),'| aa: %.4f' % (aa*100),'| kappa: %.4f' % (kappa*100))
                    writer.add_scalars(f"Epochbased/test/{iterNum}", {"loss":loss.item(),"test acc":oa},epoch)

                    # draw in tensor
                    if section['network']=='Minato':
                        if section['select_token_mode'][0]!='1':
                            show_image(name=f"matrix/linear1/{iterNum}",step=epoch,data=[model.qkvlinear1.weight.detach().cpu()])
                            show_image(name=f"matrix/linear2/{iterNum}",step=epoch,data=[model.qkvlinear2.weight.detach().cpu()])
                            if model.attn!=None and model.attn!=[] :
                                att_map=[ i.unsqueeze(0) for i in list(torch.cat(model.attn,dim=0))]
                                show_image(name=f"matrix/att_map/{iterNum}",step=epoch,data=att_map)
                            if model.v!=None and model.v!=[]:
                                v=[ i.unsqueeze(0) for i in list(torch.cat(model.v,dim=0))]
                                show_image(name=f"matrix/v/{iterNum}",step=epoch,data=v)
                        else:
                            show_image(name=f"matrix/wa1,wa2/{iterNum}",step=epoch,data=[model.wa_1.detach().cpu(),model.wa_2.detach().cpu()])
                            show_image(name=f"matrix/wb1,wb2/{iterNum}",step=epoch,data=[model.wb_1.detach().cpu(),model.wb_2.detach().cpu()]) 
                        if  section['pos_emb']!='none':
                            show_image(name=f"matrix/pos/{iterNum}",step=epoch,data=[model.pos.detach().cpu()])

                        # show_image(name=f"matrix/input_tokens/{iterNum}",step=epoch,data=[model.input_tokens])
                        # show_image(name=f"matrix/output_tokens/{iterNum}",step=epoch,data=[model.output_tokens])
                        # show_image(name=f"matrix/normed_tokens/{iterNum}",step=epoch,data=[model.normed_tokens])
                        # show_image(name=f"matrix/input_output_norm/{iterNum}",step=epoch,data=[model.input_tokens.detach().cpu(),model.output_tokens.detach().cpu(),model.normed_tokens.detach().cpu()])
                    # save the parameters in network
                    if oa > BestAcc:
                            BestAcc = oa
                            fileName = save_checkpoint(checkpointDatasetPath,configName+"_saveCheckpoint.pkl",model)
                            bestmodel = model
                    model.train()  
                    scheduler.step()



        torch.cuda.synchronize()
        end = time.time()
        print(end - start)
        Train_time = end - start

        # # load the saved parameters
        # if 'TEST' in configName:
        #     #调试不必打印保存
        #     return

        model.load_state_dict(torch.load(fileName))
        os.remove(fileName)

        model.eval()
        confusion, oa, each_acc, aa, kappa = reports(testLoader,model,datasetConfig.classNum,datasetName)
        KAPPA.append(kappa)
        OA.append(oa)
        AA.append(aa)
        ELEMENT_ACC[iterNum, :] = each_acc
        torch.save(model, checkpointPath / datasetName / f'best_model_{checkpointName}_OA={oa}_AA={aa}_Iter={iterNum}_{current_time}_{datasetName}.pt')
        recordExcel(oa,aa,kappa,checkpointName,ELEMENT_ACC)

    print("----------" + datasetName + " Training Finished -----------")
    record_output(OA, AA, KAPPA, ELEMENT_ACC,resultPath / f"{current_time}_{checkpointName}_{datasetName}")

    writer.add_hparams(
       params,
        {
            "max_OA":max(OA),
            "max_AA":max(AA),
            "max_kappa":max(KAPPA),
        })
    writer.flush()
    writer.close()
    return checkpointPath / datasetName / f'best_model_{checkpointName}_OA={oa}_AA={aa}_Iter={iterNum}_{current_time}_{datasetName}.pt'



# # Draw

# In[13]:


def padding_hsi( input_normalize,mode='mirror'):
    patch = patchsize
    height, width, band = input_normalize.shape
    padding = patch // 2
    mirror_hsi = np.zeros((height + 2 * padding, width + 2 * padding, band), dtype=float)  # padding后的图 上下左右各加padding

    mirror_hsi[padding:(padding + height), padding:(padding + width), :] = input_normalize  # 中间用原图初始化
    if mode == 'mirror':
        for i in range(padding):
            mirror_hsi[padding:(height + padding), i, :] = input_normalize[:, padding - i - 1, :]

        for i in range(padding):
            mirror_hsi[padding:(height + padding), width + padding + i, :] = input_normalize[:, width - 1 - i, :]

        for i in range(padding):
            mirror_hsi[i, :, :] = mirror_hsi[padding * 2 - i - 1, :, :]

        for i in range(padding):
            mirror_hsi[height + padding + i, :, :] = mirror_hsi[height + padding - 1 - i, :, :]
    elif mode == 'zero':
        for i in range(padding):
            mirror_hsi[padding:(height + padding), i, :] = 0

        for i in range(padding):
            mirror_hsi[padding:(height + padding), width + padding + i, :] = 0

        for i in range(padding):
            mirror_hsi[i, :, :] = 0

        for i in range(padding):
            mirror_hsi[height + padding + i, :, :] = 0

    print("**************************************************")
    print("patch is : {}".format(patch))
    print("mirror_image shape : [{0},{1},{2}]".format(mirror_hsi.shape[0], mirror_hsi.shape[1], mirror_hsi.shape[2]))
    print("**************************************************")
    return mirror_hsi
def gain_neighborhood_pixel(mirror_image, point, i, patch=5):
    x = point[i, 0]
    y = point[i, 1]
    temp_image = mirror_image[x:(x + patch), y:(y + patch), :]
    return temp_image
def true_data(mirror_image, band,  true_point, patch=11):
    x_true = np.zeros((true_point.shape[0], patch, patch, band), dtype=float)  #
    for k in range(true_point.shape[0]):
        x_true[k, :, :, :] = gain_neighborhood_pixel(mirror_image, true_point, k, patch)
    x_true = torch.from_numpy(x_true).to(torch.float32)
    x_true = x_true.permute(0,3,1,2)
    x_true = x_true.reshape(x_true.shape[0],x_true.shape[1],-1).to(torch.float32)
    return  x_true
def pred_all(model):
    # load mat
    def load_mat(modalName):
        path_img = parent_directory / "dataset" / "img" / datasetName / "{}_norm.mat".format(modalName)
        data1 = io.loadmat(path_img)

        data1 = data1['Data']
        h,w,c = data1.shape


        data1 = padding_hsi(data1,mode='mirror')
        return data1,h,w,c
    data1,h,w,c1 = load_mat(modalName=modalName1)
    data2,h,w,c2 = load_mat(modalName=modalName2)
    test_batch=int(section['test_batch'])
    number = h * w // test_batch
    total_pos_true = np.array([[i,j] for i in range(h) for j in range(w)])

    pred_all = np.empty((h*w, 1), dtype='float64')

    model.eval()
    with torch.no_grad():
        for i in range(number):
            temp_pos = total_pos_true[i * test_batch:(i + 1) * test_batch,:]
            temp_data1 = true_data(data1, c1, temp_pos, patch=patchsize,).cuda()
            temp_data2 = true_data(data2, c2, temp_pos, patch=patchsize,).cuda()
            temp2 = model(temp_data1,temp_data2)
            del temp_data1,temp_data2
            temp3 = torch.max(temp2, 1)[1].squeeze()
            del temp2
            pred_all[i * test_batch:(i + 1) * test_batch, 0] = temp3.cpu()
            del temp3

        if (i + 1) * test_batch < h * w:
            temp_pos = total_pos_true[(i + 1) * test_batch:h * w, :]
            temp_data1 = true_data(data1, c1, temp_pos, patch=patchsize,).cuda()
            temp_data2 = true_data(data2, c2, temp_pos, patch=patchsize,).cuda()
            temp2 = model(temp_data1,temp_data2)
            del temp_data1,temp_data2
            temp3 = torch.max(temp2, 1)[1].squeeze()
            del temp2
            pred_all[(i + 1) * test_batch:h * w, 0] = temp3.cpu()
            del temp3

        pred_all = np.reshape(pred_all, (h, w)) + 1
    return pred_all


# In[14]:


def draw_img(model):
    prediction_matrix = pred_all(model)
    colormap = {

    "MUUFL": colors.ListedColormap(
          ["#0000cd","#0008ff","#004dff","#0091ff","#00d4ff","#29ffce","#60ff97","#97ff60","#ceff29","#ffe600","#ffa700"]),
        "Trento": colors.ListedColormap(
           ["#0000cd","#0008ff","#004dff","#0091ff","#00d4ff","#29ffce"]),
        "Augsburg": colors.ListedColormap(
            ["#0000cd","#0008ff","#004dff","#0091ff","#00d4ff","#29ffce","#60ff97"]),
    }
    # savemat(f'{args.dataset}_matrix.mat', {'Data': prediction_matrix})
    plt.subplot(1, 1, 1)
    print(type(prediction_matrix))
    plt.imshow(prediction_matrix, cmap=colormap[datasetName])
    plt.xticks([])
    plt.yticks([])
    plt.xticks(alpha=0)
    plt.yticks(alpha=0)
    plt.axis('off')
    plt.tick_params(axis='x', width=0)
    plt.tick_params(axis='y', width=0)
    ax = plt.gca()
    ax.spines['top'].set_visible(False)
    
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['right'].set_visible(False)
    # plt.show()

    plt.tight_layout()
    # plt.legend()
    imgDatasetPath = imgPath / datasetName
    imgDatasetPath.mkdir(parents=True, exist_ok=True)

    fileName = imgDatasetPath / "{}_{}.png".format(section["network"],configName)
    plt.savefig(fileName, bbox_inches='tight', dpi=1000,pad_inches=0)
    io.savemat(imgDatasetPath / "{}_{}.mat".format(section["network"],configName),{'Data':prediction_matrix})



# In[15]:


# def test():
#     model = torch.load_state_dict(torch.load(fileName))

#             model.eval()
#             confusion, oa, each_acc, aa, kappa = reports(testLoader,model,datasetConfig.classNum,datasetName)
#             KAPPA.append(kappa)
#             OA.append(oa)
#             AA.append(aa)
#             ELEMENT_ACC[iterNum, :] = each_acc
#             torch.save(model, checkpointPath / datasetName / f'best_model_{checkpointName}_OA={oa}_AA={aa}_Iter={iterNum}_{current_time}_{datasetName}.pt')
#             recordExcel(oa,aa,kappa,checkpointName)

#         print("----------" + datasetName + " Training Finished -----------")
#         record_output(OA, AA, KAPPA, ELEMENT_ACC,resultPath / f"{current_time}_{checkpointName}_{datasetName}")
        
#         writer.add_hparams(
#            params,
#             {
#                 "max_OA":max(OA),
#                 "max_AA":max(AA),
#                 "max_kappa":max(KAPPA),
#             })
#         writer.flush()
#         writer.close()


# In[16]:


# from torchsummary import summary
# def getParams(datasetName,modelName,patchsize,band1,band2,classNum):# get patams of HCT\MFT\ExViT\Ours

#     model = getModel(modelName,band1,band2,classNum,patchsize).cuda()
#     model.eval()
#     # print(((64,datasetConfig.bandNum1, patchsize**2), (64, datasetConfig.bandNum2,patchsize**2)))
#     print(model)
#     summary(model, [(band1, patchsize**2), ( band2,patchsize**2)],batch_size = 64)


# In[17]:


from torchsummaryX import summary
def getParams(datasetName,modelName,patchsize,band1,band2,classNum):
    model = getModel(modelName,band1,band2,classNum,patchsize)
    model.eval()
    summary(model, torch.zeros((64,band1, patchsize**2)), torch.zeros((64, band2,patchsize**2)))


# In[18]:


if section['do']=='train' or section['do']=='trainAndDraw' :
    filename = train()
    print(filename)
    model = torch.load(filename)
    if section['do']=='trainAndDraw':
        # print(draw)
        draw_img(model)
elif section['do']=='draw':
    model = torch.load(checkpointDatasetPath / section['load_model'])
    draw_img(model)
elif section['do']=='params':
    datasetConfig = DatasetConfig(datasetName)
    trainLoader = datasetConfig.getTrainLoader(type = "Tr",batchsize = batchsize)
    modelNames = ["MFT","HCT","ExViT","Minato"]
    for m in modelNames:
        print(datasetName,modalName1,modalName2,m)
        getParams(datasetName=datasetName,modelName=m,patchsize=patchsize,band1=datasetConfig.bandNum1,band2=datasetConfig.bandNum2,classNum=datasetConfig.classNum)



# In[ ]:





# In[19]:


# from fastai.vision.all import *
# # from migrating_pytorch import *
# from fastai.optimizer import OptimWrapper
# import fastai.callback.schedule
# from fastai.metrics import accuracy
# from functools import partial
# def fastAi():
#     datasetConfig = DatasetConfig(datasetName)
#
#     trainLoader = datasetConfig.getTrainLoader(type = "Tr",batchsize = batchsize)
#     testLoader = datasetConfig.getTrainLoader(type = "Te",batchsize = testSizeNumber)
#     KAPPA = []
#     OA = []
#     AA = []
#     ELEMENT_ACC = np.zeros((1, datasetConfig.classNum))
#
#
#     model = Minato(datasetConfig.classNum, HSIOnly,datasetConfig.shape1,datasetConfig.shape2).cuda()
#     checkpointDatasetPath = checkpointPath / datasetName
#     checkpointDatasetPath.mkdir(parents=True, exist_ok=True)
#     checkpointDatasetPath = checkpointDatasetPath / "saveCheckpoint"
#     # optimizer = torch.optim.Adam(model.parameters(), lr=LR,weight_decay=5e-3)
#     opt_func = partial(OptimWrapper, opt=torch.optim.Adam)
#     loss_func = nn.CrossEntropyLoss()  # the target label is not one-hotted
#     # scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=50, gamma=0.9)
#
#     data = DataLoaders(trainLoader, testLoader)
#     learn = Learner(data, model, loss_func=loss_func, opt_func=opt_func, metrics=accuracy)
#
#     # train
#     # learn.fit_one_cycle(n_epoch=1, lr_max=1e-2)
#     # learn.lr_find()
#     lrs = learn.lr_find(suggest_funcs=(minimum, steep, valley, slide))
#     print(lrs)
#     # print('Learning rate with the minimum loss:', lr_min)
#     # print('Learning rate with the steepest gradient:', lr_steep)
# #     learn.save(checkpointDatasetPath, with_opt=False)
#
# #     model.load_state_dict(torch.load(str(checkpointDatasetPath)+".pth"))
# #     model.eval()
# #     confusion, oa, each_acc, aa, kappa = reports(testLoader,model,datasetConfig.classNum,datasetName)
# #     KAPPA.append(kappa)
# #     OA.append(oa)
# #     AA.append(aa)
# #     ELEMENT_ACC[0, :] = each_acc
# #     torch.save(model, checkpointPath / datasetName / f'best_model_{checkpointName}_OA={oa}_AA={aa}_Iter={0}_{current_time}_{datasetName}.pt')
# #     recordExcel(oa,aa,kappa,checkpointName)
# #     print("----------" + datasetName + " Training Finished -----------")
# #     record_output(OA, AA, KAPPA, ELEMENT_ACC,resultPath / f"{current_time}_{checkpointName}_{datasetName}")
# #     writer.add_hparams(
# #        params,
# #         {
# #             "max_OA":max(OA),
# #             "max_AA":max(AA),
# #             "max_kappa":max(KAPPA),
# #         })
# #     writer.flush()
# #     writer.close()
# if section['do']=='fastai':
#     fastAi()
#


# In[ ]:





# In[ ]:




